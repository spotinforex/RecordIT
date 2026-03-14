from utils.retry import retry
import logging
from logic.db import DatabaseConnection
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

PERIOD_MAP = {
    "today": 1,
    "7d": 7,
    "30d": 30,
    "90d": 90
}

@retry(max_attempts=3, delay=1.0, backoff=2.0, exceptions=(Exception,))
def single_complaint_retriever(complaint_id):
    try:
        db = DatabaseConnection()
        logging.info(f"Fetching complaint data for complainant {complaint_id}")
        complaint_id = complaint_id.upper()
        results = db.fetch_all('SELECT * FROM public.complaint WHERE "Complainant Code" = %s', (complaint_id,))
        db.close()
        if not results:
            logging.info(f"No records found for complainant id: {complaint_id}")
            return None
        columns = ["id", "date", "complainant_code", "cohort", "type_of_complainant",
                   "complainant_name", "phone_number", "complaint_category",
                   "communication_channel", "complainant_feedback"]
        complaint = dict(zip(columns, results[0]))
        if complaint.get("id"):
            complaint["id"] = str(complaint["id"])  
        if complaint.get("date"):
            complaint["date"] = str(complaint["date"])
        return complaint
    except Exception as e:
        logging.error(f"Error retrieving single complaint: {e}")
        raise


@retry(max_attempts=3, delay=1.0, backoff=2.0, exceptions=(Exception,))
def multiple_complaint_retriever(period=None):
    try:
        if period and period not in PERIOD_MAP:
            logging.warning(f"Invalid period: {period}")
            return None

        db = DatabaseConnection()
        logging.info(f"Retrieving complaints for period: {period or 'all'}")

        if period:
            from_date = datetime.now() - timedelta(days=PERIOD_MAP[period])
            results = db.fetch_all(
                'SELECT * FROM public.complaint WHERE date >= %s ORDER BY date DESC',
                (from_date,)
            )
        else:
            results = db.fetch_all('SELECT * FROM public.complaint ORDER BY date DESC')

        db.close()

        if not results:
            return []  

        columns = ["id", "date", "complainant_code", "cohort", "type_of_complainant",
                   "complainant_name", "phone_number", "complaint_category",
                   "communication_channel", "complainant_feedback"]

        complaints = [dict(zip(columns, row)) for row in results]
        for c in complaints:
            if c.get("id"):
                c["id"] = str(c["id"])  
            if c.get("date"):
                c["date"] = str(c["date"])
        return complaints

    except Exception as e:
        logging.error(f"Error retrieving multiple complaints: {e}")
        raise

@retry(max_attempts=3, delay=1.0, backoff=2.0, exceptions=(Exception,))
def complaints_to_excel(period=None):
    '''
    Retrieves complaints by period and converts to Excel bytes
    Args:
        period (str): "today", "7d", "30d", "90d" or None for all
    Returns:
        bytes: Excel file as bytes buffer, or None if invalid period
    '''
    try:
        complaints = multiple_complaint_retriever(period)
        if complaints is None:
            return None  # invalid period

        wb = Workbook()
        ws = wb.active
        ws.title = "Complaints"

        # Header styling
        headers = ["ID", "Date", "Complainant Code", "Cohort", "Type",
                   "Name", "Phone Number", "Category",
                   "Channel", "Feedback"]
        header_fill = PatternFill("solid", start_color="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

        # Data rows
        row_font = Font(name="Arial", size=10)
        for row_idx, complaint in enumerate(complaints, start=2):
            values = [
                complaint.get("id"),
                complaint.get("date"),
                complaint.get("complainant_code"),
                complaint.get("cohort"),
                complaint.get("type_of_complainant"),
                complaint.get("complainant_name"),
                complaint.get("phone_number"),
                complaint.get("complaint_category"),
                complaint.get("communication_channel"),
                complaint.get("complainant_feedback"),
            ]

            for col_idx, value in enumerate(values, start=1):
                # Uppercase complainant code
                if col_idx == 3 and isinstance(value, str):
                    value = value.upper()

                # Phone number as plain text
                if col_idx == 7 and value is not None:
                    value = str(value)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.number_format = '@'
                    cell.font = row_font
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill("solid", start_color="D9E1F2")
                    continue

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = row_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if row_idx % 2 == 0:
                    cell.fill = PatternFill("solid", start_color="D9E1F2")

        # Auto column widths
        col_widths = [6, 22, 18, 8, 10, 20, 15, 18, 12, 50]
        for col, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save to bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    except Exception as e:
        logging.error(f"Error generating complaints Excel: {e}")
        raise
